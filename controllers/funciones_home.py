
# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid  # Modulo de python para crear un string

from conexion.conexionBD import connectionBD  # Conexión a BD

import datetime
import re
import os

from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio


import openpyxl  # Para generar el excel
# biblioteca o modulo send_file para forzar la descarga
from flask import send_file


def procesar_form_empleado(dataForm, foto_perfil):
    # Formateando Salario
   # salario_sin_puntos = re.sub('[^0-9]+', '', dataForm['salario_empleado'])
    # convertir salario a INT
    #salario_entero = int(salario_sin_puntos)
    result_foto_perfil = procesar_imagen_perfil(foto_perfil)
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO tbl_empleados (nombre_empleado, apellido_empleado, apellido_materno, sexo_empleado, telefono_empleado, email_empleado, profesion_empleado, entrada, salida, curp, rfc, num_personal, num_plaza, categoria, fecha_ingreso, antiguedad, nss, control_asistencia, tipo_empleado, otro_empleado, calle, num, colonia, municipio, estado, cp, tipo_sangre, observaciones, foto_empleado) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"

                # Creando una tupla con los valores del INSERT
                valores = (dataForm['nombre_empleado'], dataForm['apellido_empleado'], dataForm['apellido_materno'], dataForm['sexo_empleado'],
                           dataForm['telefono_empleado'], dataForm['email_empleado'], dataForm['profesion_empleado'], dataForm['entrada'], dataForm['salida'],
                           dataForm['curp'], dataForm['rfc'], dataForm['num_personal'], dataForm['num_plaza'], dataForm['categoria'], 
                           dataForm['fecha_ingreso'], dataForm['antiguedad'], dataForm['nss'], dataForm['control_asistencia'], dataForm['tipo_empleado'], dataForm['otro_empleado'], dataForm['calle'],
                           dataForm['num'], dataForm['colonia'], dataForm['municipio'], dataForm['estado'], dataForm['cp'], dataForm['tipo_sangre'], dataForm['observaciones'], result_foto_perfil)
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_empleado: {str(e)}'


def procesar_imagen_perfil(foto):
    try:
        # Nombre original del archivo
        filename = secure_filename(foto.filename)
        extension = os.path.splitext(filename)[1]

        # Creando un string de 50 caracteres
        nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
        nombreFile = nuevoNameFile + extension

        # Construir la ruta completa de subida del archivo
        basepath = os.path.abspath(os.path.dirname(__file__))
        upload_dir = os.path.join(basepath, f'../static/fotos_empleados/')

        # Validar si existe la ruta y crearla si no existe
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            # Dando permiso a la carpeta
            os.chmod(upload_dir, 0o755)

        # Construir la ruta completa de subida del archivo
        upload_path = os.path.join(upload_dir, nombreFile)
        foto.save(upload_path)

        return nombreFile

    except Exception as e:
        print("Error al procesar archivo:", e)
        return []


# Lista de Empleados
def sql_lista_empleadosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.apellido_materno,
                        e.telefono_empleado,
                        e.entrada,
                        e.salida,
                        e.curp,
                        e.rfc,
                        e.num_personal,
                        e.num_plaza,
                        e.categoria,
                        e.fecha_ingreso,
                        e.antiguedad,
                        e.nss,
                        e.control_asistencia,
                        e.tipo_empleado,
                        e.otro_empleado,
                        e.calle,
                        e.num,
                        e.colonia,
                        e.municipio,
                        e.estado,
                        e.cp,
                        e.tipo_sangre,
                        e.observaciones,
                        e.foto_empleado,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_empleadosBD: {e}")
        return None


# Detalles del Empleado
def sql_detalles_empleadosBD(idEmpleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.apellido_materno,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado,
                        e.telefono_empleado, 
                        e.email_empleado,
                        e.profesion_empleado,
                        e.entrada,
                        e.salida,
                        e.curp,
                        e.rfc,
                        e.num_personal,
                        e.num_plaza,
                        e.categoria,
                        e.fecha_ingreso,
                        e.antiguedad,
                        e.nss,
                        e.control_asistencia,
                        e.tipo_empleado,
                        e.otro_empleado,
                        e.calle,
                        e.num,
                        e.colonia,
                        e.municipio,
                        e.estado,
                        e.cp,
                        e.tipo_sangre,
                        e.observaciones,
                        e.foto_empleado,
                        DATE_FORMAT(e.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                    FROM tbl_empleados AS e
                    WHERE id_empleado =%s
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL, (idEmpleado,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_empleadosBD: {e}")
        return None


# Funcion Empleados Informe (Reporte)
def empleadosReporte():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.apellido_materno,
                        e.email_empleado,
                        e.telefono_empleado,
                        e.profesion_empleado,
                        e.entrada,
                        e.salida,
                        e.curp,
                        e.rfc,
                        e.num_personal,
                        e.num_plaza,
                        e.categoria,
                        e.fecha_ingreso,
                        e.antiguedad,
                        e.nss,
                        e.control_asistencia,
                        e.tipo_empleado,
                        e.otro_empleado,
                        e.calle,
                        e.num,
                        e.colonia,
                        e.municipio,
                        e.estado,
                        e.cp,
                        e.tipo_sangre,
                        e.observaciones,
                        DATE_FORMAT(e.fecha_registro, '%d de %b %Y %h:%i %p') AS fecha_registro,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función empleadosReporte: {e}")
        return None


def generarReporteExcel():
    dataEmpleados = empleadosReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("Nombre", "Apellido_Paterno", "Apellido_Materno", "Sexo",
                     "Telefono", "Email", "Profesión", "Entrada", "Salida", "CURP", "RFC", "num_personal", 
                     "num_plaza", "categoria", "fecha_ingreso", "antiguedad", "NSS", "control_asistencia","Tipo_empleado","Otros", "calle", 
                     "num", "colonia", "municipio", "estado", "C.P.", "tipo_sangre", "Observaciones", "Fecha de Ingreso")

    hoja.append(cabeceraExcel)

    # Formato para números en moneda colombiana y sin decimales
    #formato_moneda_colombiana = '#,##0'

    # Agregar los registros a la hoja
    for registro in dataEmpleados:
        nombre_empleado = registro['nombre_empleado']
        apellido_empleado = registro['apellido_empleado']
        apellido_materno = registro['apellido_materno']
        sexo_empleado = registro['sexo_empleado']
        telefono_empleado = registro['telefono_empleado']
        email_empleado = registro['email_empleado']
        profesion_empleado = registro['profesion_empleado']
        entrada = registro['entrada']
        salida = registro['salida']
        #salario_empleado = registro['salario_empleado']
        curp = registro['curp']
        rfc = registro['rfc']
        num_personal = registro['num_personal']
        num_plaza = registro['num_plaza']
        categoria = registro['categoria']
        fecha_ingreso = registro['fecha_ingreso']
        antiguedad = registro['antiguedad']
        nss = registro['nss']
        control_asistencia = registro['control_asistencia']
        tipo_empleado = registro['tipo_empleado']
        otro_empleado = registro['otro_empleado']
        calle = registro['calle']
        num = registro['num']
        colonia = registro['colonia']
        municipio = registro['municipio']
        estado = registro['estado']
        cp = registro['cp']
        tipo_sangre = registro['tipo_sangre']
        observaciones = registro['observaciones']
        fecha_registro = registro['fecha_registro']

        # Agregar los valores a la hoja
        hoja.append((nombre_empleado, apellido_empleado, apellido_materno, sexo_empleado, telefono_empleado, email_empleado, profesion_empleado,
                     entrada, salida, curp, rfc, num_personal, num_plaza, categoria, fecha_ingreso, antiguedad, nss, 
                     control_asistencia, tipo_empleado, otro_empleado, calle, num, colonia, municipio, estado, cp, tipo_sangre, observaciones, fecha_registro))

        # Itera a través de las filas y aplica el formato a la columna G
      #  for fila_num in range(2, hoja.max_row + 1):
       #     columna = 8  # Columna H
        #    celda = hoja.cell(row=fila_num, column=columna)
         #   celda.number_format = formato_moneda_colombiana

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_empleados_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)


def buscarEmpleadoBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.apellido_materno,
                            e.email_empleado,
                            e.telefono_empleado,
                            e.profesion_empleado,
                            e.entrada,
                            e.salida,
                            e.curp,
                            e.rfc,
                            e.num_personal,
                            e.num_plaza,
                            e.categoria,
                            e.fecha_ingreso,
                            e.antiguedad,
                            e.nss,
                            e.control_asistencia,
                            e.tipo_empleado,
                            e.otro_empleado,
                            e.calle,
                            e.num,
                            e.colonia,
                            e.municipio,
                            e.estado,
                            e.cp,
                            e.tipo_sangre,
                            e.observaciones,
                            CASE
                                WHEN e.sexo_empleado = 1 THEN 'Masculino'
                                ELSE 'Femenino'
                            END AS sexo_empleado
                        FROM tbl_empleados AS e
                        WHERE e.nombre_empleado LIKE %s 
                        ORDER BY e.id_empleado DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoBD: {e}")
        return []


def buscarEmpleadoUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.apellido_materno,
                            e.sexo_empleado,
                            e.telefono_empleado,
                            e.email_empleado,
                            e.profesion_empleado,
                            e.entrada,
                            e.salida,
                            e.curp,
                            e.rfc,
                            e.num_personal,
                            e.num_plaza,
                            e.categoria,
                            e.fecha_ingreso,
                            e.antiguedad,
                            e.nss,
                            e.control_asistencia,
                            e.tipo_empleado,
                            e.otro_empleado,
                            e.calle,
                            e.num,
                            e.colonia,
                            e.municipio,
                            e.estado,
                            e.cp,
                            e.tipo_sangre,
                            e.observaciones,
                            e.foto_empleado
                        FROM tbl_empleados AS e
                        WHERE e.id_empleado =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                empleado = mycursor.fetchone()
                return empleado

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoUnico: {e}")
        return []


def procesar_actualizacion_form(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                nombre_empleado = data.form['nombre_empleado'] 
                apellido_empleado = data.form['apellido_empleado']
                apellido_materno = data.form['apellido_materno']
                sexo_empleado = data.form['sexo_empleado']
                telefono_empleado = data.form['telefono_empleado']
                email_empleado = data.form['email_empleado']
                profesion_empleado = data.form['profesion_empleado']
                entrada = data.form['entrada']
                salida = data.form['salida']
                curp = data.form['curp']
                rfc = data.form['rfc']
                num_personal = data.form['num_personal']
                num_plaza = data.form['num_plaza']
                categoria = data.form['categoria']
                fecha_ingreso = data.form['fecha_ingreso']
                antiguedad = data.form['antiguedad']
                nss = data.form['nss']
                control_asistencia = data.form['control_asistencia']
                tipo_empleado = data.form['tipo_empleado']
                otro_empleado = data.form['otro_empleado']
                calle = data.form['calle']
                num = data.form['num']
                colonia = data.form['colonia']
                municipio = data.form['municipio']
                estado = data.form['estado']
                cp = data.form['cp']
                tipo_sangre = data.form['tipo_sangre']
                observaciones = data.form['observaciones']
                #salario_sin_puntos = re.sub(
                # '[^0-9]+', '', data.form['salario_empleado'])
                #salario_empleado = int(salario_sin_puntos)
                id_empleado = data.form['id_empleado']

                if data.files['foto_empleado']:
                    file = data.files['foto_empleado']
                    fotoForm = procesar_imagen_perfil(file)

                    querySQL = """
                        UPDATE tbl_empleados
                        SET 
                            nombre_empleado = %s,
                            apellido_empleado = %s,
                            apellido_materno = %s,
                            sexo_empleado = %s,
                            telefono_empleado = %s,
                            email_empleado = %s,
                            profesion_empleado = %s,
                            entrada = %s,
                            salida = %s,
                            curp = %s,
                            rfc = %s,
                            num_personal = %s,
                            num_plaza = %s,
                            categoria = %s,
                            fecha_ingreso = %s,
                            antiguedad = %s,
                            nss = %s,
                            control_asistencia = %s,
                            tipo_empleado = %s,
                            otro_empleado = %s,
                            calle = %s,
                            num = %s,
                            colonia = %s,
                            municipio = %s,
                            estado = %s,
                            cp = %s,
                            tipo_sangre = %s,
                            observaciones = %s,
                            foto_empleado = %s
                        WHERE id_empleado = %s
                    """
                    values = (nombre_empleado, apellido_empleado, apellido_materno, sexo_empleado,
                              telefono_empleado, email_empleado, profesion_empleado, entrada, salida, curp, 
                              rfc, num_personal, num_plaza, categoria, fecha_ingreso, antiguedad, 
                              nss, control_asistencia, tipo_empleado, otro_empleado, calle, num, colonia, municipio, estado, cp, 
                              tipo_sangre, observaciones, fotoForm, id_empleado)
                else:
                    querySQL = """
                        UPDATE tbl_empleados
                        SET 
                            nombre_empleado = %s,
                            apellido_empleado = %s,
                            apellido_materno = %s,
                            sexo_empleado = %s,
                            telefono_empleado = %s,
                            email_empleado = %s,
                            profesion_empleado = %s,
                            entrada = %s,
                            salida = %s,
                            curp = %s,
                            rfc = %s,
                            num_personal = %s,
                            num_plaza = %s,
                            categoria = %s,
                            fecha_ingreso = %s,
                            antiguedad = %s,
                            nss = %s,
                            control_asistencia = %s,
                            tipo_empleado = %s,
                            otro_empleado = %s,
                            calle = %s,
                            num = %s,
                            colonia = %s,
                            municipio = %s,
                            estado = %s,
                            cp = %s,
                            tipo_sangre = %s,
                            observaciones = %s
                        WHERE id_empleado = %s
                    """
                    values = (nombre_empleado, apellido_empleado, apellido_materno, sexo_empleado,
                              telefono_empleado, email_empleado, profesion_empleado, entrada, salida, 
                              curp, rfc,  num_personal, num_plaza, categoria, fecha_ingreso,
                              antiguedad, nss, control_asistencia, tipo_empleado, otro_empleado, calle, num, colonia, 
                              municipio, estado, cp, tipo_sangre, observaciones, id_empleado)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {e}")
        return None


# Lista de Usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id, name_surname, email_user, created_user FROM users"
                cursor.execute(querySQL,)
                usuariosBD = cursor.fetchall()
        return usuariosBD
    except Exception as e:
        print(f"Error en lista_usuariosBD : {e}")
        return []


# Eliminar uEmpleado
def eliminarEmpleado(id_empleado, foto_empleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_empleados WHERE id_empleado=%s"
                cursor.execute(querySQL, (id_empleado,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

                if resultado_eliminar:
                    # Eliminadon foto_empleado desde el directorio
                    basepath = path.dirname(__file__)
                    url_File = path.join(
                        basepath, '../static/fotos_empleados', foto_empleado)

                    if path.exists(url_File):
                        remove(url_File)  # Borrar foto desde la carpeta

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarEmpleado : {e}")
        return []


# Eliminar usuario
def eliminarUsuario(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM users WHERE id=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []
